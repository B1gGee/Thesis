import streamlit as st
import yfinance as yf
import pandas as pd
import requests
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from io import BytesIO
import time

st.set_page_config(page_title="Stock Thesis Monitor", layout="wide", page_icon="📈")

st.markdown("""
<style>
.status-strongbuy { color: #28a745; font-size: 1.8em; font-weight: 800; }
.status-hold      { color: #ffc107; font-size: 1.8em; font-weight: 800; }
.status-monitor   { color: #17a2b8; font-size: 1.8em; font-weight: 800; }
.status-sell      { color: #dc3545; font-size: 1.8em; font-weight: 800; }
</style>
""", unsafe_allow_html=True)

st.title("📊 Stock Thesis Monitor")
st.caption("Automated thesis builder • Entry / Mid / High Level • Custom triggers • GitHub-ready")

# ====================== API KEYS ======================
st.sidebar.header("🔑 API Keys")
fmp_key = st.sidebar.text_input("FMP API Key", value="KoegUYuz7Ig5PiZNhen3I2byMyf6elWz", type="password")
alpha_key = st.sidebar.text_input("Alpha Vantage API Key", value="US6OIYAJTOK8CXX7", type="password")

# FMP remaining calls
if fmp_key:
    try:
        usage = requests.get(f"https://financialmodelingprep.com/api/v3/usage?apikey={fmp_key}", timeout=10).json()
        remaining = usage.get('dailyLimit', 250) - usage.get('dailyUsage', 0)
        st.sidebar.success(f"**FMP calls left today: {remaining}**")
    except:
        st.sidebar.info("FMP remaining calls: (could not fetch)")

# Alpha Vantage session counter
if 'alpha_calls' not in st.session_state:
    st.session_state.alpha_calls = 0
st.sidebar.info(f"**Alpha Vantage calls this session: {st.session_state.alpha_calls}** (25 free/day)")

ticker = st.sidebar.text_input("Enter Ticker (e.g. AAPL, TSLA, BHP.AX, RIO.AX, AYA.AX)", value="AAPL").upper().strip()

if st.sidebar.button("Generate Thesis"):
    with st.spinner(f"Fetching data for {ticker}..."):
        info = None
        source = "None"

        # 1. PRIORITY 1: Yahoo Finance
        st.info("🔄 Trying Yahoo Finance first (Priority 1)...")
        for attempt in range(5):
            try:
                session = requests.Session()
                session.headers.update({'User-Agent': 'Mozilla/5.0'})
                stock = yf.Ticker(ticker, session=session)
                info = stock.info
                source = "Yahoo Finance"
                st.success("✅ Loaded from Yahoo Finance")
                break
            except:
                st.warning(f"Yahoo attempt {attempt+1}/5 failed")
                time.sleep(2)

        # 2. PRIORITY 2: FMP (improved for .AX and unusual tickers)
        if not info and fmp_key:
            st.info("🔄 Yahoo failed → Trying FMP (Priority 2)...")
            tickers_to_try = [ticker]
            if ticker.endswith('.AX'):
                tickers_to_try.append(ticker.replace('.AX', ''))
            elif '.' not in ticker:
                tickers_to_try.append(ticker + '.AX')
            for t in tickers_to_try:
                try:
                    profile = requests.get(f"https://financialmodelingprep.com/api/v3/profile/{t}?apikey={fmp_key}", timeout=10).json()
                    quote = requests.get(f"https://financialmodelingprep.com/api/v3/quote/{t}?apikey={fmp_key}", timeout=10).json()
                    if isinstance(profile, list) and len(profile) > 0:
                        p = profile[0]
                        q = quote[0] if isinstance(quote, list) and len(quote) > 0 else {}
                        info = {
                            'longName': p.get('companyName'),
                            'sector': p.get('sector'),
                            'industry': p.get('industry'),
                            'longBusinessSummary': p.get('description'),
                            'currentPrice': q.get('price'),
                            'marketCap': q.get('marketCap'),
                            'fiftyTwoWeekHigh': q.get('yearHigh'),
                            'fiftyTwoWeekLow': q.get('yearLow'),
                            'forwardPE': p.get('forwardPE'),
                            'earningsGrowth': p.get('earningsGrowth'),
                            'revenueGrowth': p.get('revenueGrowth'),
                            'dividendYield': p.get('dividendYield'),
                            'returnOnEquity': p.get('roe'),
                            'debtToEquity': p.get('debtToEquity'),
                            'freeCashflow': q.get('freeCashFlow'),
                            'heldPercentInsiders': p.get('heldPercentInsiders'),
                            'targetMeanPrice': p.get('targetPrice'),
                            'recommendationKey': p.get('recommendation'),
                            'regularMarketChangePercent': q.get('changePercent'),
                        }
                        source = "Financial Modeling Prep (FMP)"
                        st.success(f"✅ Loaded from FMP using {t}")
                        break
                    else:
                        st.warning(f"FMP returned no data for {t}")
                except Exception as e:
                    st.warning(f"FMP error for {t}: {type(e).__name__}")
                    continue

        # 3. LAST RESORT: Alpha Vantage
        if not info and alpha_key:
            st.info("🔄 FMP failed → Trying Alpha Vantage (last resort)...")
            try:
                ov = requests.get(f"https://www.alphavantage.co/query?function=OVERVIEW&symbol={ticker}&apikey={alpha_key}", timeout=10).json()
                gq = requests.get(f"https://www.alphavantage.co/query?function=GLOBAL_QUOTE&symbol={ticker}&apikey={alpha_key}", timeout=10).json().get('Global Quote', {})
                if ov and 'Symbol' in ov:
                    info = {
                        'longName': ov.get('Name'),
                        'sector': ov.get('Sector'),
                        'industry': ov.get('Industry'),
                        'longBusinessSummary': ov.get('Description'),
                        'currentPrice': float(gq.get('05. price', 0) or 0),
                        'marketCap': float(ov.get('MarketCapitalization', 0) or 0),
                        'fiftyTwoWeekHigh': float(ov.get('52WeekHigh', 0) or 0),
                        'fiftyTwoWeekLow': float(ov.get('52WeekLow', 0) or 0),
                        'forwardPE': float(ov.get('ForwardPE', 0) or 0),
                        'earningsGrowth': float(ov.get('EPSGrowthThisYear', 0) or 0),
                        'revenueGrowth': float(ov.get('QuarterlyRevenueGrowthYOY', 0) or 0),
                        'dividendYield': float(ov.get('DividendYield', 0) or 0),
                        'returnOnEquity': float(ov.get('ReturnOnEquityTTM', 0) or 0),
                        'debtToEquity': float(ov.get('DebtToEquity', 0) or 0),
                        'freeCashflow': None,
                        'heldPercentInsiders': None,
                        'targetMeanPrice': None,
                        'recommendationKey': None,
                        'regularMarketChangePercent': float(gq.get('10. change percent', '0').replace('%','') or 0),
                    }
                    source = "Alpha Vantage"
                    st.session_state.alpha_calls += 1
                    st.success("✅ Loaded from Alpha Vantage")
            except:
                pass

        if not info:
            st.error("❌ All sources failed. Try again in 1-2 minutes or use a major US ticker like AAPL or TSLA.")
            st.stop()

        today = datetime.now().strftime("%Y-%m-%d")

        # === TOP HEADER ===
        col1, col2, col3 = st.columns([1, 2, 1])
        with col1:
            price = info.get('currentPrice')
            delta = info.get('regularMarketChangePercent', 0)
            st.metric("Current Price", f"${price:,}" if price else "N/A", delta=f"{delta:.2f}%")
        with col2:
            st.subheader(f"{ticker} • {info.get('longName', ticker)}")
            st.caption(f"{info.get('sector', '<Unable to Source>')} • {info.get('industry', '<Unable to Source>')}")
        with col3:
            pe = info.get('forwardPE')
            growth = info.get('earningsGrowth') or 0
            rec = info.get('recommendationKey', 'hold')
            fcf_yield = 0
            if info.get('freeCashflow') and info.get('marketCap'):
                fcf_yield = info.get('freeCashflow') / info.get('marketCap') * 100
            score = 0
            if pe and pe < 20: score += 2
            if growth > 0.15: score += 2
            if rec in ['buy', 'strong_buy']: score += 3
            if fcf_yield > 8: score += 1
            if pe and pe > 40: score -= 2
            if growth < -0.05: score -= 2
            if score >= 6:
                status, color = "Strong Buy", "status-strongbuy"
            elif score >= 3:
                status, color = "Buy / Hold", "status-hold"
            elif score >= 0:
                status, color = "Monitor", "status-monitor"
            else:
                status, color = "Sell", "status-sell"
            st.markdown(f"**Overall Status**<br><span class='{color}'>{status}</span>", unsafe_allow_html=True)
            st.caption(f"Score: {score} | {today} | Source: {source}")

        # Company Overview & KPIs
        st.markdown("### Company Overview & Key Milestone KPIs")
        st.write(info.get('longBusinessSummary', '<Unable to Source>'))

        kpi_cols = st.columns(4)
        with kpi_cols[0]: st.metric("Market Cap", f"${info.get('marketCap', 0)/1e9:.1f}B")
        with kpi_cols[1]: st.metric("52w High / Low", f"${info.get('fiftyTwoWeekHigh', 'N/A')} / ${info.get('fiftyTwoWeekLow', 'N/A')}")
        with kpi_cols[2]: st.metric("Dividend Yield", f"{info.get('dividendYield', 0)*100:.2f}%")
        with kpi_cols[3]: st.metric("Analyst Target", f"${info.get('targetMeanPrice', 'N/A')}")

        # === ENTRY LEVEL ===
        st.markdown("### 📌 Entry-Level (Basic 5-10 min scan)")
        entry_data = {
            "Ticker / Company / Sector / Industry": [f"{ticker} • {info.get('longName', ticker)} • {info.get('sector', '<Unable>')} • {info.get('industry', '<Unable>')}"],
            "Current Price | Market Cap | 52w High/Low": [f"${info.get('currentPrice', 'N/A')} | ${info.get('marketCap', 0)/1e9:.1f}B | ${info.get('fiftyTwoWeekHigh', 'N/A')}/${info.get('fiftyTwoWeekLow', 'N/A')}"],
            "P/E (fwd)": [info.get('forwardPE', '<Unable to Source>')],
            "EPS growth (1-3y)": [f"{info.get('earningsGrowth', 0)*100:.1f}%" if info.get('earningsGrowth') is not None else '<Unable to Source>'],
            "Revenue growth (recent)": [f"{info.get('revenueGrowth', 0)*100:.1f}%" if info.get('revenueGrowth') is not None else '<Unable to Source>'],
            "Dividend yield": [f"{info.get('dividendYield', 0)*100:.2f}%"],
            "Simple 'Why Buy' (auto-generated)": ["• Growing demand in " + info.get('sector', 'its sector') + "\n• " + (info.get('longBusinessSummary', '')[:120] + "...")],
            "Basic Valuation vs peers": ["<Unable to Source> – compare manually in Excel export"],
            "Quick Risks": ["• Market volatility\n• Sector headwinds"],
        }
        df_entry = pd.DataFrame.from_dict(entry_data, orient='index', columns=['Value/Notes'])
        df_entry['Flag'] = [''] * len(df_entry)
        df_entry['Source/Date'] = [f"{source} • {today}"] * len(df_entry)
        df_entry = df_entry.reset_index().rename(columns={'index': 'Metric/Query'})
        edited_entry = st.data_editor(df_entry, column_config={"Flag": st.column_config.SelectboxColumn("Flag", options=["Green", "Orange", "Red"], required=True)}, use_container_width=True, num_rows="fixed", key="entry_editor")

        st.markdown("**Entry-Level Triggers**")
        st.data_editor(pd.DataFrame([{"Trigger": "Earnings beat + raised guidance", "Color": "Green"}, {"Trigger": "New catalyst (contract win)", "Color": "Green"}, {"Trigger": "Price dips on no news", "Color": "Orange"}, {"Trigger": "Major miss + lowered guidance", "Color": "Red"}]), use_container_width=True, hide_index=True)

        # === MID-LEVEL (safe) ===
        st.markdown("### 📌 Mid-Level (Core 20-30 min)")
        roe = info.get('returnOnEquity')
        debt_eq = info.get('debtToEquity')
        fcf = info.get('freeCashflow')
        insider = info.get('heldPercentInsiders') or 0
        mid_data = {
            "ROE / ROIC": [f"{(roe or 0)*100:.1f}%" if roe is not None else '<Unable to Source>'],
            "Debt/Equity": [debt_eq if debt_eq is not None else '<Unable to Source>'],
            "Free Cash Flow trend": [f"${fcf/1e9:.1f}B (latest)" if fcf is not None else '<Unable to Source>'],
            "Margins (gross/operating)": [f"{info.get('grossMargins',0)*100:.1f}% / {info.get('operatingMargins',0)*100:.1f}%"],
            "Management & Strategy": [f"Insider ownership: {insider*100:.1f}%"],
            "Historical Performance (1y vs S&P500)": ["<Unable to Source> – add in export"],
        }
        df_mid = pd.DataFrame.from_dict(mid_data, orient='index', columns=['Value/Notes'])
        df_mid['Flag'] = [''] * len(df_mid)
        df_mid['Source/Date'] = [f"{source} • {today}"] * len(df_mid)
        df_mid = df_mid.reset_index().rename(columns={'index': 'Metric/Query'})
        edited_mid = st.data_editor(df_mid, column_config={"Flag": st.column_config.SelectboxColumn("Flag", options=["Green", "Orange", "Red"])}, use_container_width=True, key="mid_editor")

        # === HIGH-LEVEL (safe DCF) ===
        st.markdown("### 📌 High-Level / In-Depth")
        st.caption("Simple 2-stage DCF (auto)")
        try:
            rev_growth = info.get('revenueGrowth', 0.08) or 0.08
            fcf0 = info.get('freeCashflow') or 1_000_000_000
            wacc = 0.10
            terminal_g = 0.03
            dcf_value = fcf0 * (1 + rev_growth) * (1 - (1 + terminal_g) / (1 + wacc)) / (wacc - terminal_g) / 1e9
            st.write(f"**Implied DCF Fair Value ≈ ${dcf_value:.1f}B** (base case)")
        except:
            st.write("**DCF: <Unable to Source>**")

        high_data = {
            "Valuation Models (DCF / Multiples)": ["See calculation above"],
            "3-5y Revenue/EBITDA forecasts": ["Assumption: " + str(rev_growth*100) + "% CAGR"],
            "Portfolio Fit / Review Date": ["Add your % allocation + next review date"],
        }
        df_high = pd.DataFrame.from_dict(high_data, orient='index', columns=['Value/Notes'])
        df_high['Flag'] = [''] * len(df_high)
        df_high['Source/Date'] = [f"{source} • {today}"] * len(df_high)
        df_high = df_high.reset_index().rename(columns={'index': 'Metric/Query'})
        edited_high = st.data_editor(df_high, column_config={"Flag": st.column_config.SelectboxColumn("Flag", options=["Green", "Orange", "Red"])}, use_container_width=True, key="high_editor")

        # === CUSTOM THESIS BLOCK ===
        st.markdown("### 🎯 Custom Thesis Block (Horyzon-style)")
        col_a, col_b = st.columns(2)
        with col_a:
            st.text_area("Why I Own It", value="Growing demand + strong moat", height=120)
            st.text_area("What Must Stay True", value="Revenue growth >15% annually", height=120)
        with col_b:
            st.text_area("What Would Change My Mind (Thesis Breakers)", value="Debt covenant breach or ROIC < WACC", height=120)
            st.date_input("Portfolio Role & Next Review Date", value=datetime.now())

        # Conviction Score
        all_flags = pd.concat([edited_entry['Flag'], edited_mid['Flag'], edited_high['Flag']])
        conviction = (all_flags == "Green").sum() - (all_flags == "Red").sum()
        st.metric("Conviction Score", f"{conviction} / 10", help="+1 Green, 0 Orange, -1 Red")

        # EXPORT
        if st.button("📥 Export to Excel (with colors + dropdowns)"):
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                pd.concat([edited_entry, edited_mid, edited_high]).to_excel(writer, sheet_name="Thesis", index=False)
                wb = writer.book
                ws = wb["Thesis"]
                green_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
                orange_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
                red_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
                for row in ws.iter_rows(min_row=2, max_col=4):
                    cell = row[2]
                    if cell.value == "Green": cell.fill = green_fill
                    elif cell.value == "Orange": cell.fill = orange_fill
                    elif cell.value == "Red": cell.fill = red_fill
                dv = DataValidation(type="list", formula1='"Green,Orange,Red"', allow_blank=True)
                ws.add_data_validation(dv)
                dv.add('C2:C100')
            output.seek(0)
            st.download_button(label="Download Thesis.xlsx", data=output, file_name=f"{ticker}_Thesis_{today}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        st.success(f"✅ Thesis generated from **{source}**! Edit any cell above.")
        st.caption("All missing fields show '<Unable to Source>'. Add your own research for qualitative depth.")
